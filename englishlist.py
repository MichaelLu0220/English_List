import os
import openpyxl
import time
import tkinter as tk
import win32gui
import random

def show_message(message):
    message_window = tk.Toplevel()
    x,y = get_screen_center()
    message_window.geometry(f'160x50+{x-80}+{y-25}')
    message_label = tk.Label(message_window, text=message)
    message_label.pack()
    message_window.after(3000, message_window.destroy)

def get_screen_center():
    root = tk.Tk()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.destroy()  
    center_x = screen_width // 2
    center_y = screen_height // 2
    return center_x, center_y

class EnglishListApp:
    def __init__(self):
        self.wb = openpyxl.load_workbook('English_words_list.xlsx')
        self.s1 = self.wb['file1']

        self.which_one = 0
        self.change_press = 0
        self.lock = ''
        self.mod = 0
        self.switch = None
        self.bottom_display = False
        self.inactive_time = 3000
        self.hide_buttons_id = None
        self.dynamic_width = False
        self.max_text_length = 0
        self.showing_buttons = False

        self.windows = tk.Tk()
        self.windows.lift()
        self.windows.attributes('-topmost', True)
        self.windows.title('EnglishList')
        self.windows.geometry('500x50')
        self.windows.resizable(False, False)

        self.ntime = 3000
        self.text_content = tk.StringVar()
        self.text_content.set('')
        self.label = tk.Label(self.windows, textvariable=self.text_content)

        self.text1 = tk.Text(self.windows, width=31, height=1, font=('Times New Roman', 17, 'bold'))
        self.text1.tag_configure("center", justify='center')
        self.text1.grid(row=0, column=4, padx=10, pady=1, sticky='nsew')

        self.mybutton = tk.Button(self.windows, text='change', command=self.button_event)
        self.bt_plus = tk.Button(self.windows, text='+1', command=self.plus_)
        self.bt_min = tk.Button(self.windows, text='-1', command=self.min_)
        self.toggle_width_button = tk.Button(self.windows, text='TW', command=self.toggle_width_mode)

        self.option_button = tk.Button(self.windows, text='option', command=self.show_buttons)
        self.option_button.grid(row=0, column=1, padx=5, pady=5)
        self.label.grid(row=1, column=4, padx=0, pady=1)

        self.windows.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.calculate_max_text_length()
        self.update_text_width()
        self.refreshText()

    def is_empty_row(self, row):
        return all(cell.value is None for cell in row)

    def get_next_valid_row(self, start_row, max_row, step=1):
        row = start_row
        while row >= 1 and row <= max_row:
            cells = list(self.s1.iter_rows(min_row=row, max_row=row, min_col=1, max_col=3))[0]
            if not self.is_empty_row(cells):
                return row
            row += step
        return None

    def calculate_max_text_length(self):
        max_length = 0
        for row in self.s1.iter_rows(min_col=1, max_col=3):
            text = '  '.join(str(cell.value) for cell in row if cell.value is not None)
            max_length = max(max_length, len(text))
        self.max_text_length = max_length

    def refreshText(self):
        max_row = int(self.s1.max_row)

        if self.mod == 1:
            self.which_one = random.randint(1, max_row)
        else:
            self.which_one += 1
            if self.which_one > max_row:
                self.which_one = 1

        next_row = self.get_next_valid_row(self.which_one, max_row, step=1)
        if next_row is None:
            self.which_one = 1
            next_row = self.get_next_valid_row(self.which_one, max_row, step=1)

        if next_row:
            self.which_one = next_row
            self.text1.delete(0.0, tk.END)
            v = self.s1.iter_rows(min_row=self.which_one, min_col=1, max_col=3, max_row=self.which_one)
            a = ''
            for k in v:
                for j in k:
                    a += (str(j.value) if j.value is not None else "") + "  "
            self.text1.insert(tk.INSERT, a)
            self.text1.tag_add("center", 1.0, "end")
            self.text1.update()
            self.update_text_width()
        self.switch = self.windows.after(self.ntime, self.refreshText) 

    def update_text_width(self):
        if self.dynamic_width:
            text1_width = len(self.text1.get("1.0", "end-1c")) + 5
        else:
            text1_width = self.max_text_length + 2

        self.text1.config(width=text1_width)

        if text1_width < 10:
            text1_width += 2

        if self.bottom_display:
            if self.showing_buttons:
                self.windows.geometry(f'{(text1_width) * 13 + 160}x60')
            else:
                self.windows.geometry(f'{(text1_width) * 13 + 60}x60')

        else:
            if self.showing_buttons:
                self.windows.geometry(f'{(text1_width) * 13 + 160}x50')
            else:
                self.windows.geometry(f'{(text1_width) * 13 + 60}x50')

    def button_event(self):
        if self.change_press == 0:
            self.change_press = 1
            self.lock += '123'
            if self.lock == '123456789123':
                if self.mod == 0:
                    show_message("\nChange Mod to Random")
                    self.mod = 1
                else:
                    show_message("\nChange Mod to Normal")
                    self.mod = 0
                self.lock = ''
            elif self.lock == '123' or self.lock == '123123':
                self.lock = '123'
            else:
                self.lock = ''
        else:
            self.change_press = 0
        self.stop()
        self.reset_timer()

    def stop(self):
        if self.change_press == 1:
            self.text1.after_cancel(self.switch)
            self.text1.delete(0.0, tk.END)

            max_row = int(self.s1.max_row)
            self.which_one += 1
            if self.which_one > max_row:
                self.which_one = 1

            next_row = self.get_next_valid_row(self.which_one, max_row, step=1)
            if next_row is None:
                self.which_one = 1
                next_row = self.get_next_valid_row(self.which_one, max_row, step=1)

            if next_row:
                self.which_one = next_row
                v = self.s1.iter_rows(min_row=self.which_one, min_col=1, max_col=3, max_row=self.which_one)
                a = ''
                for k in v:
                    for j in k:
                        a += (str(j.value) if j.value is not None else "") + "  "
                self.text1.insert(tk.INSERT, a)
                self.text1.tag_add("center", 1.0, "end")

            self.change_press = 0
            self.update_text_width()
            self.stop()
        else:
            self.switch = self.windows.after(self.ntime, self.refreshText)

    def plus_(self):
        self.lock += '456'
        if '456456' in self.lock:
            self.lock = ''
        self.ntime += 1000
        self.show()
        self.windows.after(1000, self.clear)
        self.reset_timer()

    def min_(self):
        self.lock += '789'
        if '789789' in self.lock:
            self.lock = ''
        self.ntime -= 1000
        self.show()
        self.windows.after(1000, self.clear)
        self.reset_timer()

    def show(self):
        self.bottom_display = True
        if self.ntime < 1000:
            self.ntime = 1000
            self.text_content.set("Too fast")
        else:
            self.text_content.set("Time set: " + str(self.ntime / 1000) + " sec")
        self.update_text_width()

    def clear(self):
        self.text_content.set('')
        self.bottom_display = False
        self.update_text_width()

    def on_closing(self):
        if self.switch is not None:
            self.windows.after_cancel(self.switch)
        self.windows.quit()

    def show_buttons(self):
        self.option_button.grid_forget()
        self.mybutton.grid(row=0, column=1, padx=5, pady=5)
        self.bt_plus.grid(row=0, column=2, padx=5, pady=5)
        self.bt_min.grid(row=0, column=3, padx=5, pady=5)
        self.toggle_width_button.grid(row=0, column=0, padx=2, pady=2)
        self.showing_buttons = True
        self.reset_timer()
        self.update_text_width()

    def hide_buttons(self):
        if self.hide_buttons_id is not None:
            self.windows.after_cancel(self.hide_buttons_id)
        self.option_button.grid(row=0, column=1, padx=5, pady=5)
        self.mybutton.grid_forget()
        self.bt_plus.grid_forget()
        self.bt_min.grid_forget()
        self.toggle_width_button.grid_forget()
        self.hide_buttons_id = None
        self.showing_buttons = False
        self.update_text_width()

    def reset_timer(self):
        if self.hide_buttons_id is not None:
            self.windows.after_cancel(self.hide_buttons_id)
        self.hide_buttons_id = self.windows.after(self.inactive_time, self.hide_buttons)

    def toggle_width_mode(self):
        self.dynamic_width = not self.dynamic_width
        if not self.dynamic_width:
            self.calculate_max_text_length()
        self.update_text_width()

app = EnglishListApp()
app.windows.mainloop()

